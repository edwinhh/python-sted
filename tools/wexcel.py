import openpyxl

from openpyxl import Workbook

from openpyxl import load_workbook

#写
# book = openpyxl.Workbook()
# sheet = book.active #默认的sheet
# #sheet2 = book.get_sheet_by_name('sheet1')
# # sheet.append( ['id','username','password','error_count'])
# # sheet.append( [1,'wyj','123456',0])
# # sheet.append( [2,'wyj','123456'])
# sheet['a1'] = 'id' #指定行列
# sheet['b1'] = 'username'#
# sheet.cell(3,1,'1')#指定行和列
# book.save('user.xlsx')

# print(sheet.cell(1,1).value)
# print(sheet['a1'].value)
# print(list(sheet.rows)) #所有行的数据
# l = []
# print(list(sheet.columns)) #所有列

# for row in sheet.rows:
#     t = []
#     for col in row:
#         t.append(col.value)
#     l.append(t)
# print(l)

# print(sheet[1:10])#第几行到第几行


exfile = "D:\python_sted\day8\用例模板.xlsx"


def get_data(filename, sheet_name, title=1,line="all"):


    data = []  # 申明list

    workbook_ = openpyxl.load_workbook(filename)  # 导入工作表

    sheetnames = workbook_.get_sheet_names()  # 获得表单名字

    # sheet = workbook_.get_sheet_by_name(sheetnames[sheet_name]) #从工作表中提取某一表单

    sheet = workbook_.get_sheet_by_name(sheet_name)

    print("Work Sheet Rows:", sheet.max_row)

    #print(sheet.max_column)
    if line=="all":
        line=sheet.max_row + 1

    for rowNum in range(title, line):
        data_ = []
        for colNum in range(1, sheet.max_column + 1):
            data_.append(sheet.cell(row=rowNum, column=colNum).value)
        data.append(data_)




        #
        # else:
        #
        #     url_ = []
        #
        #     for colNum in range(1, sheet.max_column + 1):
        #         url_.append(sheet.cell(row=rowNum, column=colNum).value)
        #
        #     data.append(url_)



    return data




def wexcel(name,sheet_name,data):
    wb = Workbook()

    sheet = wb.create_sheet(sheet_name, index=0)
    sheet = wb.active

    for i in data:
        sheet.append(i)  # 这一步就可以做到将12345插入到一行中。
    wb.save(name)

# name="reslut.xlsx"
# sheet_name="1"
# a = get_data(exfile, "1")
# print(a)
# wexcel(name,sheet_name,a)
